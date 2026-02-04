# fosa/utils_import_geo.py
from django.db import transaction
from openpyxl import load_workbook
from .models import Wilaya, Moughataa, Commune

def _norm(s):
    if s is None:
        return ""
    return " ".join(str(s).strip().split()).lower()

def import_geo_from_xlsx(filepath: str, update_if_exists: bool = False):
    """
    Lit un .xlsx et importe Wilayas/Moughataas/Communes.
    Colonnes attendues (insensibles à la casse) :
      - wilaya, wilaya_code (optionnel)
      - moughataa, moughataa_code (optionnel)
      - commune, commune_code (optionnel)
    Renvoie un dict de stats.
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = None
    rownum = 0

    stats = {
        "wilayas":   {"created": 0, "updated": 0, "skipped": 0, "errors": 0},
        "moughataas":{"created": 0, "updated": 0, "skipped": 0, "errors": 0},
        "communes":  {"created": 0, "updated": 0, "skipped": 0, "errors": 0},
        "rows": 0,
    }

    @transaction.atomic
    def _run():
        nonlocal headers, rownum
        for row in ws.iter_rows(values_only=True):
            rownum += 1
            if rownum == 1:
                headers = [_norm(h) for h in row]
                continue
            if not headers:
                raise ValueError("Entêtes manquantes (ligne 1).")
            stats["rows"] += 1

            data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            wilaya_name    = (data.get("wilaya") or "") .strip()
            wilaya_code    = (data.get("wilaya_code") or None)
            moughataa_name = (data.get("moughataa") or "") .strip()
            moughataa_code = (data.get("moughataa_code") or None)
            commune_name   = (data.get("commune") or "") .strip()
            commune_code   = (data.get("commune_code") or None)

            # ----- Wilaya -----
            try:
                w = None
                if wilaya_code:
                    w = Wilaya.objects.filter(code=str(wilaya_code).strip()).first()
                if not w and wilaya_name:
                    w = Wilaya.objects.filter(nom__iexact=wilaya_name).first()

                if w:
                    if update_if_exists:
                        changed = False
                        if wilaya_code and w.code != str(wilaya_code).strip():
                            w.code = str(wilaya_code).strip(); changed = True
                        if wilaya_name and w.nom != wilaya_name:
                            w.nom = wilaya_name; changed = True
                        if changed:
                            w.save(update_fields=["nom", "code"])
                            stats["wilayas"]["updated"] += 1
                        else:
                            stats["wilayas"]["skipped"] += 1
                    else:
                        stats["wilayas"]["skipped"] += 1
                else:
                    if not wilaya_name:
                        stats["wilayas"]["errors"] += 1
                        # Impossible d’aller plus loin sans wilaya
                        continue
                    w = Wilaya.objects.create(
                        nom=wilaya_name,
                        code=str(wilaya_code).strip() if wilaya_code else None
                    )
                    stats["wilayas"]["created"] += 1
            except Exception:
                stats["wilayas"]["errors"] += 1
                continue

            # ----- Moughataa -----
            m = None
            if moughataa_name:
                try:
                    if moughataa_code:
                        m = Moughataa.objects.filter(wilaya=w, code=str(moughataa_code).strip()).first()
                    if not m:
                        m = Moughataa.objects.filter(wilaya=w, nom__iexact=moughataa_name).first()

                    if m:
                        if update_if_exists:
                            changed = False
                            if moughataa_code and m.code != str(moughataa_code).strip():
                                m.code = str(moughataa_code).strip(); changed = True
                            if m.nom != moughataa_name:
                                m.nom = moughataa_name; changed = True
                            if changed:
                                m.save(update_fields=["nom", "code"])
                                stats["moughataas"]["updated"] += 1
                            else:
                                stats["moughataas"]["skipped"] += 1
                        else:
                            stats["moughataas"]["skipped"] += 1
                    else:
                        m = Moughataa.objects.create(
                            wilaya=w, nom=moughataa_name,
                            code=str(moughataa_code).strip() if moughataa_code else None
                        )
                        stats["moughataas"]["created"] += 1
                except Exception:
                    stats["moughataas"]["errors"] += 1
                    m = None  # pas de parent => pas de commune

            # ----- Commune -----
            if commune_name:
                try:
                    if not moughataa_name:
                        stats["communes"]["errors"] += 1
                        continue
                    # retrouve la moughataa (créée/trouvée ci-dessus)
                    if not m:
                        q = Moughataa.objects.filter(wilaya=w, nom__iexact=moughataa_name)
                        if moughataa_code:
                            q = q | Moughataa.objects.filter(wilaya=w, code=str(moughataa_code).strip())
                        m = q.first()
                        if not m:
                            stats["communes"]["errors"] += 1
                            continue

                    c = None
                    if commune_code:
                        c = Commune.objects.filter(moughataa=m, code=str(commune_code).strip()).first()
                    if not c:
                        c = Commune.objects.filter(moughataa=m, nom__iexact=commune_name).first()

                    if c:
                        if update_if_exists:
                            changed = False
                            if commune_code and c.code != str(commune_code).strip():
                                c.code = str(commune_code).strip(); changed = True
                            if c.nom != commune_name:
                                c.nom = commune_name; changed = True
                            if changed:
                                c.save(update_fields=["nom", "code"])
                                stats["communes"]["updated"] += 1
                            else:
                                stats["communes"]["skipped"] += 1
                        else:
                            stats["communes"]["skipped"] += 1
                    else:
                        Commune.objects.create(
                            moughataa=m, nom=commune_name,
                            code=str(commune_code).strip() if commune_code else None
                        )
                        stats["communes"]["created"] += 1
                except Exception:
                    stats["communes"]["errors"] += 1
                    continue

    _run()
    return stats
