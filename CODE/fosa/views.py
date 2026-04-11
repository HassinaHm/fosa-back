from accounts.permissions import CustomModelPermissions, FOSARolePermission
from rest_framework import viewsets, permissions, serializers, generics, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from import_export import resources
from tablib import Dataset
from django.contrib.auth.models import User
from .models import FOSA, FOSAHistory
import logging
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.db.models import Q
from import_export import resources
from .models import FOSA, TypeStructure
import json

from .serializers import (
    FOSASerializer, WilayaSerializer, MoughataaSerializer, CommuneSerializer,
    MaladieSerializer, MaladieReportSerializer, TypeStructureSerializer,
    NormePersonnelSerializer, NormeServiceSerializer, NormeMaterielSerializer,
    PersonnelStructureSerializer, ServiceStructureSerializer, MaterielStructureSerializer,  # ADD THESE
    FOSAHistorySerializer
)


logger = logging.getLogger(__name__)



# fosa/views_geo.py
from rest_framework import viewsets, filters
# from django_filters.rest_framework import DjangoFilterBackend
from .models import Wilaya, Moughataa, Commune
from .serializers import  FOSASerializer, WilayaSerializer, MoughataaSerializer, CommuneSerializer
from accounts.permissions import CustomModelPermissions  # si tu veux verrouiller CRUD



class WilayaViewSet(viewsets.ModelViewSet):
    serializer_class = WilayaSerializer

    def get_queryset(self):
        qs = Wilaya.objects.all().order_by("nom")
        user = self.request.user

        if not user.is_authenticated:
            return qs.none()

        if getattr(user, "is_superuser", False):
            return qs

        role = getattr(getattr(user, "role", None), "nom", None)

        if role == "Administrateur national":
            return qs

        if role == "gestionnaire régional":
            wilaya_ids = list(user.wilayas.values_list("id", flat=True))
            wilaya_noms = list(user.wilayas.values_list("nom", flat=True))
            return qs.filter(Q(id__in=wilaya_ids) | Q(nom__in=wilaya_noms))

        if role == "gestionnaire local":
            wilaya_ids = list(user.wilayas.values_list("id", flat=True))
            wilaya_noms = list(user.wilayas.values_list("nom", flat=True))
            return qs.filter(Q(id__in=wilaya_ids) | Q(nom__in=wilaya_noms))

        return qs.none()



class MoughataaViewSet(viewsets.ModelViewSet):
    serializer_class = MoughataaSerializer
    filterset_fields = ["wilaya"]
    search_fields = ["nom", "code", "wilaya__nom"]

    def get_queryset(self):
        qs = Moughataa.objects.select_related("wilaya").all()
        user = self.request.user

        if not user.is_authenticated:
            return qs.none()

        if getattr(user, "is_superuser", False):
            return qs

        role = getattr(getattr(user, "role", None), "nom", None)

        if role == "Administrateur national":
            return qs

        if role == "gestionnaire régional":
            wilaya_ids = list(user.wilayas.values_list("id", flat=True))
            wilaya_noms = list(user.wilayas.values_list("nom", flat=True))
            return qs.filter(Q(wilaya_id__in=wilaya_ids) | Q(wilaya__nom__in=wilaya_noms))

        if role == "gestionnaire local":
            return qs.filter(
                Q(id=user.moughataa_fk_id) |
                Q(nom=getattr(user.moughataa_fk, "nom", None))
            )

        return qs.none()



class CommuneViewSet(viewsets.ModelViewSet):
    serializer_class = CommuneSerializer
    filterset_fields = ["moughataa"]
    search_fields = ["nom", "code", "moughataa__nom", "moughataa__wilaya__nom"]

    def get_queryset(self):
        qs = Commune.objects.select_related("moughataa", "moughataa__wilaya").all()
        user = self.request.user

        if not user.is_authenticated:
            return qs.none()

        if getattr(user, "is_superuser", False):
            return qs

        role = getattr(getattr(user, "role", None), "nom", None)

        if role == "Administrateur national":
            return qs

        if role == "gestionnaire régional":
            wilaya_ids = list(user.wilayas.values_list("id", flat=True))
            wilaya_noms = list(user.wilayas.values_list("nom", flat=True))
            return qs.filter(
                Q(moughataa__wilaya_id__in=wilaya_ids) |
                Q(moughataa__wilaya__nom__in=wilaya_noms)
            )

        if role == "gestionnaire local":
            q = qs.filter(
                Q(moughataa_id=user.moughataa_fk_id) |
                Q(moughataa__nom=getattr(user.moughataa_fk, "nom", None))
            )
            if user.commune_fk_id:
                q = q.filter(id=user.commune_fk_id)
            return q

        return qs.none()




# fosa/views_import.py
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings

from .utils_import_geo import import_geo_from_xlsx
from accounts.permissions import CustomModelPermissions  

class GeoImportView(APIView):
    """
    POST multipart/form-data:
      - file: .xlsx
      - update_if_exists: "true" / "false"
    """
    parser_classes = [MultiPartParser]
    permission_classes = [permissions.IsAuthenticated ,CustomModelPermissions, FOSARolePermission]

    def post(self, request):
        f = request.FILES.get("file")
        update_if_exists = (request.data.get("update_if_exists") or "").lower() == "true"
        if not f:
            return Response({"detail": "Aucun fichier reçu (clé 'file')."}, status=400)

        temp_path = default_storage.save(f"tmp/geo_import/{f.name}", ContentFile(f.read()))
        absolute = default_storage.path(temp_path)

        try:
            stats = import_geo_from_xlsx(absolute, update_if_exists=update_if_exists)
            return Response({"status": "ok", "update_if_exists": update_if_exists, "stats": stats})
        except Exception as e:
            return Response({"status": "error", "detail": str(e)}, status=400)
        finally:
            # nettoyage best-effort
            try:
                default_storage.delete(temp_path)
            except Exception:
                pass



# # Serializer principal
# class FOSASerializer(serializers.ModelSerializer):
#     class Meta:
#         model = FOSA
#         fields = [
#             'nom_fr', 'nom_ar', 'type', 'code_etablissement',
#             'longitude', 'latitude', 'coordonnees', 'adresse',
#             'responsable', 'commune', 'moughataa', 'wilaya',
#             'departement', 'is_public'
#         ]
#         read_only_fields = ['code_etablissement', 'coordonnees', 'adresse', 'is_public']

#     def validate(self, data):
#         errors = {}
#         if not data.get('nom_fr'):
#             errors['nom_fr'] = "Ce champ est obligatoire"
#         if not data.get('type'):
#             errors['type'] = "Ce champ est obligatoire"
#         if not data.get('wilaya') or not data.get('moughataa') or not data.get('commune'):
#             errors['localisation'] = "Wilaya, Moughataa et commune sont obligatoires"
#         if errors:
#             raise serializers.ValidationError(errors)
#         return data


# Import/Export


class FOSAResource(resources.ModelResource):
    class Meta:
        model = FOSA
        import_id_fields = ['code_etablissement']
        fields = (
            'code_etablissement',
            'structure',               # nouveau nom principal
            'nom_fr',
            'nom_ar',
            'type',
            'type_structure',          # nouveau champ (code ou libellé)
            'departement',
            'responsable',
            'adresse',
            'commune',
            'moughataa',
            'wilaya',
            'coordonnees',
            'latitude',
            'longitude',
            'is_public',
            # nouveaux champs de StructureSante
            'etat',
            'etat_batiment',
            'cloture',
            'electricite',
            'internet',
            'eau',
            'cdf',
            'equipement',
            'date_de_construction',
            'fosa_reference',
            'fosa_plus_proche',
            'prestation_service',
            'service_manquant',
            'besoins',
            'pourcentage_activite',
            'observation',
            'bailleur',
            'source_file',
        )
        skip_unchanged = True
        report_skipped = True
        use_transactions = False

    TYPE_MAPPING = {
        'poste de santé': 'PS',
        'PS': 'PS',
        'centre de santé': 'CS',
        'CS': 'CS',
        'CH': 'CH',
        'direction régionale de santé': 'DRS',
        'DRS': 'DRS',
        'direction centrale': 'DAF',
        'DAF': 'DAF',
        'FOND': 'FOND',
        'autres': 'AUTRE',
    }

    # ------------------------------------------------------------
    # Helpers de nettoyage
    # ------------------------------------------------------------
    def parse_bool(self, value):
        if value is None:
            return None
        s = str(value).strip().lower()
        if s in ('1', 'true', 'vrai', 'oui', 'y', 'yes'):
            return True
        if s in ('0', 'false', 'faux', 'non', 'n', 'no'):
            return False
        return None

    def parse_list(self, value):
        if value is None:
            return []
        s = str(value).strip()
        if not s:
            return []
        if s.startswith('[') and s.endswith(']'):
            try:
                obj = json.loads(s)
                if isinstance(obj, list):
                    return obj
            except:
                pass
        if ';' in s:
            return [x.strip() for x in s.split(';') if x.strip()]
        if ',' in s:
            return [x.strip() for x in s.split(',') if x.strip()]
        return [s]

    def clean_type(self, raw_value):
        if not raw_value:
            return "AUTRE"
        value = str(raw_value).strip().replace("é", "e")
        return self.TYPE_MAPPING.get(value, "AUTRE")

    def clean_coordinates(self, row):
        lat = row.get('latitude', '')
        lon = row.get('longitude', '')
        if lat in ('', ',', 'nan', 'none', None, 'None'):
            row['latitude'] = None
        else:
            try:
                if isinstance(lat, str):
                    row['latitude'] = float(lat.strip())
                else:
                    row['latitude'] = float(lat)
            except (ValueError, TypeError):
                raise ValueError(f"Latitude invalide : {lat}")

        if lon in ('', ',', 'nan', 'none', None, 'None'):
            row['longitude'] = None
        else:
            try:
                if isinstance(lon, str):
                    row['longitude'] = float(lon.strip())
                else:
                    row['longitude'] = float(lon)
            except (ValueError, TypeError):
                raise ValueError(f"Longitude invalide : {lon}")

    # ------------------------------------------------------------
    # Avant importation de chaque ligne
    # ------------------------------------------------------------
    def before_import_row(self, row, **kwargs):
        # Type
        row['type'] = self.clean_type(row.get('type'))

        # Coordonnées
        self.clean_coordinates(row)

        # Booléens
        for bf in ['cloture', 'electricite', 'internet', 'eau', 'cdf']:
            row[bf] = self.parse_bool(row.get(bf))

        # Listes JSON
        row['prestation_service'] = self.parse_list(row.get('prestation_service'))
        row['service_manquant'] = self.parse_list(row.get('service_manquant'))

        # Remplir structure si vide mais nom_fr présent 
        if not row.get('structure') and row.get('nom_fr'):
            row['structure'] = row['nom_fr']

        # Validation obligatoire
        if not row.get('structure') and not row.get('nom_fr') and not row.get('nom_ar'):
            raise ValueError("Il faut au moins structure, nom_fr ou nom_ar")
        for field in ['commune', 'moughataa', 'wilaya']:
            if not row.get(field):
                raise ValueError(f"Le champ {field} est obligatoire")

    # ------------------------------------------------------------
    # Avant sauvegarde de l'instance
    # ------------------------------------------------------------
    def before_save_instance(self, instance, *args, **kwargs):
        # Public / privé
        instance.is_public = instance.type in [
            'PS', 'CS', 'CH', 'Poste de Santé', 'Centre de Santé', 'Centre hospitalier'
        ]

        # Résolution du type_structure (par code ou libellé)
        ts_val = getattr(instance, 'type_structure', None)
        if ts_val and not isinstance(ts_val, TypeStructure):
            ts = TypeStructure.objects.filter(code=ts_val).first() \
                  or TypeStructure.objects.filter(libelle=ts_val).first()
            instance.type_structure = ts

        # Remplir structure si manquant (par nom_fr)
        if not instance.structure and instance.nom_fr:
            instance.structure = instance.nom_fr

    # ------------------------------------------------------------
    # Récupération de l'instance existante pour update
    # ------------------------------------------------------------
    def get_instance(self, instance_loader, row):
        try:
            code = row.get('code_etablissement')
            if code:
                return self._meta.model.objects.get(code_etablissement=code)
        except self._meta.model.DoesNotExist:
            return None
        return None
    
    
    

class FOSAViewSet(viewsets.ModelViewSet):
    serializer_class = FOSASerializer
    queryset = FOSA.objects.select_related(
        "wilaya_fk", "moughataa_fk", "commune_fk", "type_structure"
    ).all()
    lookup_field = 'code_etablissement'

    permission_classes = [permissions.IsAuthenticated, CustomModelPermissions, FOSARolePermission]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'wilaya': ['exact'],
        'moughataa': ['exact'],
        'commune': ['exact'],
        'wilaya_fk': ['exact'],
        'moughataa_fk': ['exact'],
        'commune_fk': ['exact'],
        'type': ['exact'],
        'type_structure': ['exact'],
        'is_public': ['exact'],
        'etat': ['exact'],
    }
    search_fields = ['code_etablissement', 'structure', 'nom_fr', 'nom_ar', 'responsable']
    ordering_fields = ['code_etablissement', 'structure', 'type', 'is_public', 'etat']

    # ------------------------------------------------------------
    # Filtrage par rôle
    # ------------------------------------------------------------
    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user

        if not user.is_authenticated:
            return qs.filter(is_public=True)

        if user.is_superuser:
            return qs

        role = getattr(getattr(user, "role", None), "nom", None)
        if role == "Administrateur national":
            return qs

        if role == "gestionnaire régional":
            wilaya_ids = list(user.wilayas.values_list("id", flat=True))
            return qs.filter(wilaya_fk_id__in=wilaya_ids)

        if role == "gestionnaire local":
            if user.commune_fk_id:
                return qs.filter(commune_fk_id=user.commune_fk_id)
            if user.moughataa_fk_id:
                return qs.filter(moughataa_fk_id=user.moughataa_fk_id)
            return qs.none()

        return qs.filter(is_public=True)

    # ------------------------------------------------------------
    # Historique
    # ------------------------------------------------------------
    def perform_create(self, serializer):
        instance = serializer.save()
        self._create_history(instance, 'CREATE', {})

    def perform_update(self, serializer):
        instance = self.get_object()
        old_data = FOSASerializer(instance).data.copy()
        updated_instance = serializer.save()
        new_data = FOSASerializer(updated_instance).data.copy()
        changes = self._generate_diff(old_data, new_data)
        self._create_history(updated_instance, 'UPDATE', changes)

    def perform_destroy(self, instance):
        self._create_history(instance, 'DELETE', {})
        instance.delete()

    def _generate_diff(self, old, new):
        return {f: [old[f], new[f]] for f in old.keys() if old.get(f) != new.get(f)}

    def _create_history(self, instance, action, changes):
        FOSAHistory.objects.create(
            fosa=instance,
            user=self.request.user if self.request.user.is_authenticated else None,
            action=action,
            changes=changes
        )

    # ------------------------------------------------------------
    # Import / Export
    # ------------------------------------------------------------
    @action(detail=False, methods=['post'])
    def import_data(self, request):
        if 'file' not in request.FILES:
            return Response({"error": "Aucun fichier fourni"}, status=status.HTTP_400_BAD_REQUEST)
        file = request.FILES['file']
        if not file.name.lower().endswith(('.xlsx', '.xls', '.csv')):
            return Response({"error": "Formats acceptés: .xlsx, .xls, .csv"}, status=400)

        dataset = Dataset()
        try:
            if file.name.lower().endswith('.csv'):
                imported_data = dataset.load(file.read().decode('utf-8'), format='csv')
            elif file.name.lower().endswith('.xlsx'):
                imported_data = dataset.load(file.read(), format='xlsx')
            else:
                imported_data = dataset.load(file.read(), format='xls')
        except Exception as e:
            return Response({"status": "error", "error": f"Lecture fichier: {e}"}, status=400)

        resource = FOSAResource()
        result = resource.import_data(dataset, dry_run=False, raise_errors=False)
        return Response({
            "status": "success",
            "imported": result.totals.get('new', 0),
            "updated": result.totals.get('update', 0),
            "skipped": result.totals.get('skipped', 0),
            "total": len(imported_data)
        })

    @action(detail=False, methods=['get'])
    def export_data(self, request):
        resource = FOSAResource()
        dataset = resource.export()
        resp = HttpResponse(dataset.xlsx, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="fosas_export.xlsx"'
        return resp
    # ------------------------------------------------------------
    # Actions pour les normes (personnel, services, matériel)
    # ------------------------------------------------------------
    @action(detail=True, methods=["get"])
    def personnels(self, request, pk=None):
        fosa = self.get_object()
        qs = fosa.personnels.all().order_by("intitule_poste")
        return Response(PersonnelStructureSerializer(qs, many=True).data)

    @action(detail=True, methods=["post"], url_path="personnels/upsert")
    def personnels_upsert(self, request, pk=None):
        fosa = self.get_object()
        items = request.data if isinstance(request.data, list) else [request.data]
        saved = []
        for it in items:
            intitule = it.get("intitule_poste")
            nombre = it.get("nombre_reel", 0)
            if not intitule:
                return Response({"detail": "intitule_poste manquant"}, status=400)
            obj, _ = PersonnelStructure.objects.update_or_create(
                structure=fosa,
                intitule_poste=intitule,
                defaults={"nombre_reel": nombre},
            )
            saved.append(PersonnelStructureSerializer(obj).data)
        return Response(saved, status=200)

    @action(detail=True, methods=["get"])
    def services(self, request, pk=None):
        fosa = self.get_object()
        qs = fosa.services.all().order_by("nom_service")
        return Response(ServiceStructureSerializer(qs, many=True).data)

    @action(detail=True, methods=["post"], url_path="services/upsert")
    def services_upsert(self, request, pk=None):
        fosa = self.get_object()
        items = request.data if isinstance(request.data, list) else [request.data]
        saved = []
        for it in items:
            nom = it.get("nom_service")
            dispo = bool(it.get("disponible", False))
            if not nom:
                return Response({"detail": "nom_service manquant"}, status=400)
            obj, _ = ServiceStructure.objects.update_or_create(
                structure=fosa,
                nom_service=nom,
                defaults={"disponible": dispo},
            )
            saved.append(ServiceStructureSerializer(obj).data)
        return Response(saved, status=200)

    @action(detail=True, methods=["get"])
    def materiels(self, request, pk=None):
        fosa = self.get_object()
        qs = fosa.materiels.all().order_by("nom_materiel")
        return Response(MaterielStructureSerializer(qs, many=True).data)

    @action(detail=True, methods=["post"], url_path="materiels/upsert")
    def materiels_upsert(self, request, pk=None):
        fosa = self.get_object()
        items = request.data if isinstance(request.data, list) else [request.data]
        saved = []
        for it in items:
            nom = it.get("nom_materiel")
            qte = it.get("quantite_reelle", 0)
            if not nom:
                return Response({"detail": "nom_materiel manquant"}, status=400)
            obj, _ = MaterielStructure.objects.update_or_create(
                structure=fosa,
                nom_materiel=nom,
                defaults={"quantite_reelle": qte},
            )
            saved.append(MaterielStructureSerializer(obj).data)
        return Response(saved, status=200)
    @action(detail=False, methods=["get"], url_path="conformity-report")
    def conformity_report(self, request):
        """
        Calculate conformity percentage for each structure by comparing:
        - PersonnelStructure vs NormePersonnel
        - ServiceStructure vs NormeService
        - MaterielStructure vs NormeMateriel
        """
        fosas = self.get_queryset()
        conformity_data = []

        for fosa in fosas:
            if not fosa.type_structure:
                conformity_data.append({
                    "code_etablissement": fosa.code_etablissement,
                    "structure": fosa.structure or fosa.nom_fr,
                    "type": fosa.type,
                    "conformity_percentage": 0,
                    "message": "Type de structure non défini",
                    "details": {
                        "personnel": {"met": 0, "total": 0},
                        "services": {"met": 0, "total": 0},
                        "materiel": {"met": 0, "total": 0},
                    }
                })
                continue

            # ✅ PERSONNEL: Compare PersonnelStructure with NormePersonnel
            norme_personnel = NormePersonnel.objects.filter(type_structure=fosa.type_structure)
            actual_personnel = PersonnelStructure.objects.filter(structure=fosa)
            
            personnel_met = 0
            for norme in norme_personnel:
                actual = actual_personnel.filter(intitule_poste=norme.intitule_poste).first()
                if actual and actual.nombre_reel >= norme.nombre_minimal:
                    personnel_met += 1

            # ✅ SERVICES: Compare ServiceStructure with NormeService
            norme_services = NormeService.objects.filter(type_structure=fosa.type_structure, obligatoire=True)
            actual_services = ServiceStructure.objects.filter(structure=fosa)
            
            services_met = 0
            for norme in norme_services:
                actual = actual_services.filter(nom_service=norme.nom_service, disponible=True).first()
                if actual:
                    services_met += 1

            # ✅ MATERIEL: Compare MaterielStructure with NormeMateriel
            norme_materiel = NormeMateriel.objects.filter(type_structure=fosa.type_structure)
            actual_materiel = MaterielStructure.objects.filter(structure=fosa)
            
            materiel_met = 0
            for norme in norme_materiel:
                actual = actual_materiel.filter(nom_materiel=norme.nom_materiel).first()
                if actual and actual.quantite_reelle >= norme.quantite_minimale:
                    materiel_met += 1

            # ✅ Calculate total conformity percentage
            total_normes = norme_personnel.count() + norme_services.count() + norme_materiel.count()
            total_met = personnel_met + services_met + materiel_met

            if total_normes == 0:
                conformity_percentage = 0
                message = "Aucune norme définie"
            else:
                conformity_percentage = int((total_met / total_normes) * 100)
                message = f"{total_met}/{total_normes} normes respectées"

            conformity_data.append({
                "code_etablissement": fosa.code_etablissement,
                "structure": fosa.structure or fosa.nom_fr,
                "type": fosa.type,
                "wilaya": fosa.wilaya_fk.nom or fosa.wilaya,
                "moughataa": fosa.moughataa_fk.nom or fosa.moughataa,
                "conformity_percentage": conformity_percentage,
                "message": message,
                "details": {
                    "personnel": {
                        "met": personnel_met,
                        "total": norme_personnel.count(),
                    },
                    "services": {
                        "met": services_met,
                        "total": norme_services.count(),
                    },
                    "materiel": {
                        "met": materiel_met,
                        "total": norme_materiel.count(),
                    },
                }
            })

        return Response(conformity_data)
    
    
    
    
    
    
    
    
    
    
    
# Vue Historique
class FOSAHistorySerializer(serializers.ModelSerializer):
    fosa_code_etablissement = serializers.ReadOnlyField(source='fosa.code_etablissement')
    fosa_nom_fr = serializers.ReadOnlyField(source='fosa.nom_fr')
    fosa_nom_ar = serializers.ReadOnlyField(source='fosa.nom_ar')
    username = serializers.ReadOnlyField(source='user.email')

    class Meta:
        model = FOSAHistory
        fields = ['fosa_code_etablissement', 'fosa_nom_fr', 'fosa_nom_ar',
                  'username', 'action', 'changes', 'timestamp']



class FOSAHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = FOSAHistorySerializer
    queryset = FOSAHistory.objects.select_related('fosa', 'user').all()
    permission_classes = [permissions.IsAuthenticated ,CustomModelPermissions, FOSARolePermission]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user

        if not user.is_authenticated:
            return qs.filter(fosa__is_public=True)

        if getattr(user, "is_superuser", False):
            return qs

        role = getattr(getattr(user, "role", None), "nom", None)

        # --- Administrateur national ---
        if role == "Administrateur national":
            return qs

        # --- Gestionnaire régional ---
        if role == "gestionnaire régional":
            wilaya_ids = list(user.wilayas.values_list("id", flat=True))
            wilaya_noms = list(user.wilayas.values_list("nom", flat=True))

            return qs.filter(
                Q(fosa__wilaya_fk_id__in=wilaya_ids) |
                Q(fosa__wilaya__in=wilaya_noms)
            )

        # --- Gestionnaire local ---
        if role == "gestionnaire local":
            wilaya_noms = list(user.wilayas.values_list("nom", flat=True))

            q = qs.filter(
                Q(fosa__moughataa_fk_id=user.moughataa_fk_id) |
                Q(fosa__moughataa=user.moughataa.nom, fosa__wilaya__in=wilaya_noms)
            )



            return q

        # --- Utilisateurs publics ---
        return qs.filter(fosa__is_public=True)

import csv, io, json
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser,JSONParser, FormParser
from rest_framework.response import Response
from rest_framework import permissions

def to_bool(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("1", "true", "vrai", "oui", "y", "yes"):
        return True
    if s in ("0", "false", "faux", "non", "n", "no"):
        return False
    return None

def to_list(v):
    if v is None:
        return []
    if isinstance(v, list):
        return v
    s = str(v).strip()
    if not s:
        return []

    if (s.startswith("[") and s.endswith("]")) or (s.startswith("{") and s.endswith("}")):
        try:
            obj = json.loads(s)
            return obj if isinstance(obj, list) else []
        except Exception:
            pass

    if ";" in s:
        return [x.strip() for x in s.split(";") if x.strip()]
    if "," in s:
        return [x.strip() for x in s.split(",") if x.strip()]

    return [s]

# class StructureImportView(APIView):
#     parser_classes = [MultiPartParser]
#     permission_classes = [permissions.IsAuthenticated]  # (garde simple pour tester)

#     def post(self, request):
#         f = request.FILES.get("file")
#         if not f:
#             return Response({"detail": "Aucun fichier reçu (clé 'file')"}, status=400)

#         raw = f.read()
#         try:
#             text = raw.decode("utf-8")
#         except UnicodeDecodeError:
#             text = raw.decode("ISO-8859-1", errors="replace")

#         # IMPORTANT: si ton CSV est généré par Excel FR => souvent ; sinon mets delimiter=","
#         reader = csv.DictReader(io.StringIO(text), delimiter=";")

#         allowed = {
#             "code","structure","etat","nom_ar","coordonnee_gps","responsable","etat_batiment",
#             "date_de_construction","cloture","electricite","internet","eau","cdf","equipement",
#             "fosa_reference","fosa_plus_proche","besoins","pourcentage_activite",
#             "observation","bailleur","source_file",
#         }

#         created, updated, skipped = 0, 0, 0
#         errors = []

#         with transaction.atomic():
#             for i, row in enumerate(reader, start=2):
#                 try:
#                     wilaya_name = (row.get("wilaya") or "").strip()
#                     moughataa_name = (row.get("moughataa") or "").strip()
#                     commune_name = (row.get("commune") or "").strip()

#                     wilaya = Wilaya.objects.filter(nom__iexact=wilaya_name).first() if wilaya_name else None
#                     moughataa = None
#                     commune = None

#                     if wilaya and moughataa_name:
#                         moughataa = Moughataa.objects.filter(wilaya=wilaya, nom__iexact=moughataa_name).first()

#                     if moughataa and commune_name:
#                         commune = Commune.objects.filter(moughataa=moughataa, nom__iexact=commune_name).first()

#                     ts = None
#                     t = (row.get("type") or "").strip()
#                     if t:
#                         ts = TypeStructure.objects.filter(code__iexact=t).first() or TypeStructure.objects.filter(libelle__iexact=t).first()

#                     data = {k: row.get(k) for k in allowed if k in row}

#                     for bfield in ("cloture", "electricite", "internet", "eau", "cdf"):
#                         if bfield in data:
#                             data[bfield] = to_bool(data[bfield])

#                     data["prestation_service"] = to_list(row.get("prestation_service"))
#                     data["service_manquant"] = to_list(row.get("service_manquant"))

#                     data["wilaya_fk"] = wilaya
#                     data["moughataa_fk"] = moughataa
#                     data["commune_fk"] = commune
#                     data["type_structure"] = ts

#                     code = (row.get("code") or "").strip()

#                     if code:
#                         obj, is_created = StructureSante.objects.update_or_create(
#                             code=code,
#                             defaults=data
#                         )
#                     else:
#                         obj = StructureSante.objects.create(**data)
#                         is_created = True

#                     created += 1 if is_created else 0
#                     updated += 0 if is_created else 1

#                 except Exception as e:
#                     skipped += 1
#                     errors.append({"line": i, "error": str(e)})

#         return Response({
#             "status": "ok",
#             "created": created,
#             "updated": updated,
#             "skipped": skipped,
#             "errors": errors[:50],
#         })

from .models import Maladie, MaladieReport
from .serializers import MaladieSerializer, MaladieReportSerializer

class MaladieViewSet(viewsets.ModelViewSet):
    queryset = Maladie.objects.all().order_by("name")
    serializer_class = MaladieSerializer
    parser_classes = [JSONParser, FormParser, MultiPartParser]
    permission_classes = [permissions.IsAuthenticated ,CustomModelPermissions, FOSARolePermission]

from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import io

class MaladieReportViewSet(viewsets.ModelViewSet):
    queryset = MaladieReport.objects.select_related("wilaya", "moughataa", "maladie").all()
    serializer_class = MaladieReportSerializer
    parser_classes = [JSONParser, FormParser, MultiPartParser]
    permission_classes = [permissions.IsAuthenticated, CustomModelPermissions, FOSARolePermission]

    def get_queryset(self):
        qs = super().get_queryset()

        date = self.request.query_params.get("date")
        date_start = self.request.query_params.get("date_start")
        date_end = self.request.query_params.get("date_end")
        wilaya = self.request.query_params.get("wilaya")
        moughataa = self.request.query_params.get("moughataa")
        maladie = self.request.query_params.get("maladie")

        if date:
            qs = qs.filter(date=date)
        if date_start and date_end:
            qs = qs.filter(date__range=[date_start, date_end])

        if wilaya:
            qs = qs.filter(wilaya_id=wilaya)
        if moughataa:
            qs = qs.filter(moughataa_id=moughataa)
        if maladie:
            qs = qs.filter(maladie_id=maladie)

        return qs.order_by("date", "wilaya_id", "moughataa_id", "maladie_id")

    @action(detail=False, methods=["post"], url_path="upsert")
    def upsert(self, request):
        """
        Upsert par (date, wilaya, moughataa, disease).
        Si existe => update (remplacer)
        Sinon => create
        """
        key_fields = ["date", "wilaya", "moughataa", "maladie"]
        missing = [f for f in key_fields if f not in request.data]
        if missing:
            return Response({"detail": f"Champs manquants: {missing}"}, status=400)

        obj = MaladieReport.objects.filter(
            date=request.data["date"],
            wilaya_id=request.data["wilaya"],
            moughataa_id=request.data["moughataa"],
            maladie_id=request.data["maladie"],
        ).first()

        if obj:
            ser = self.get_serializer(obj, data=request.data, partial=False)
            ser.is_valid(raise_exception=True)
            ser.save()
            return Response(ser.data, status=status.HTTP_200_OK)

        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data, status=status.HTTP_201_CREATED)

    # ✅ NEW EXPORT ENDPOINT
    @action(detail=False, methods=["get"], url_path="export-weekly")
    def export_weekly(self, request):
        """
        Export MaladieReport as weekly Excel report
        Query params:
        - date_start: YYYY-MM-DD (Monday of week)
        - date_end: YYYY-MM-DD (Sunday of week)
        """
        date_start = request.query_params.get("date_start")
        date_end = request.query_params.get("date_end")

        if not date_start or not date_end:
            return Response(
                {"detail": "date_start and date_end are required (format: YYYY-MM-DD)"},
                status=400
            )

        try:
            start = datetime.strptime(date_start, "%Y-%m-%d").date()
            end = datetime.strptime(date_end, "%Y-%m-%d").date()
        except ValueError:
            return Response({"detail": "Invalid date format. Use YYYY-MM-DD"}, status=400)

        # Get week number
        week_num = start.isocalendar()[1]

        # Fetch reports for this week
        reports = MaladieReport.objects.filter(
            date__range=[start, end]
        ).select_related("wilaya", "moughataa", "maladie").order_by(
            "wilaya__nom", "moughataa__nom", "maladie__name"
        )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport Hebdomadaire"

        # ✅ Styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # ✅ Title Row
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"NOTIFICATION DES MALADIES ET EVENEMENTS"
        title_cell.font = title_font
        title_cell.alignment = center_align

        # ✅ Week Info Row
        ws.merge_cells("A2:H2")
        week_cell = ws["A2"]
        week_cell.value = f"Sem. Épid. N° : {week_num:02d}  du {start.strftime('%d/%m/%Y')} au {end.strftime('%d/%m/%Y')}"
        week_cell.font = Font(bold=True, size=10)
        week_cell.alignment = center_align

        # ✅ Statistics (Placeholder)
        row = 4
        stats = [
            ("Nombre de rapports attendus des Moughataas", len(set(reports.values_list("moughataa_id", flat=True)))),
            ("Nombre de rapports reçus des Moughataas", len(set(reports.values_list("moughataa_id", flat=True)))),
            ("Nombre de rapports reçus à temps des Moughataas", len(set(reports.values_list("moughataa_id", flat=True)))),
        ]

        for stat_label, stat_value in stats:
            ws[f"A{row}"] = stat_label
            ws[f"B{row}"] = stat_value
            ws[f"C{row}"] = f"{100}%"
            row += 1

        # ✅ Headers
        row = 8
        headers = ["Wilaya", "Moughataa", "Maladie", "Cas Suspects", "Décès", "Cas Prélevés", "Cas Testés", "Cas Confirmés"]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # ✅ Data Rows
        row = 9
        for report in reports:
            ws.cell(row=row, column=1).value = report.wilaya.nom
            ws.cell(row=row, column=2).value = report.moughataa.nom
            ws.cell(row=row, column=3).value = report.maladie.name
            ws.cell(row=row, column=4).value = report.cas_suspects or 0
            ws.cell(row=row, column=5).value = report.deces or 0
            ws.cell(row=row, column=6).value = report.cas_preleves or 0
            ws.cell(row=row, column=7).value = report.cas_testes or 0
            ws.cell(row=row, column=8).value = report.cas_confirmes or 0

            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = center_align

            row += 1

        # ✅ Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30
        for col in ["D", "E", "F", "G", "H"]:
            ws.column_dimensions[col].width = 15

        # ✅ Generate file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Rapport_Epid_Sem_{week_num:02d}_{start.strftime('%Y%m%d')}_au_{end.strftime('%Y%m%d')}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"
        return response
    
    
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response

from .models import (
    TypeStructure,
    NormePersonnel, NormeService, NormeMateriel,
    # StructureSante,
    PersonnelStructure, ServiceStructure, MaterielStructure
)


class TypeStructureViewSet(viewsets.ModelViewSet):
    queryset = TypeStructure.objects.all().order_by("libelle")
    serializer_class = TypeStructureSerializer

class NormePersonnelViewSet(viewsets.ModelViewSet):
    queryset = NormePersonnel.objects.all()
    serializer_class = NormePersonnelSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        type_structure = self.request.query_params.get("type_structure")
        # ✅ FIX: Only filter if type_structure is not null/None
        if type_structure and type_structure != "null":
            try:
                qs = qs.filter(type_structure_id=int(type_structure))
            except (ValueError, TypeError):
                pass
        return qs.order_by("intitule_poste")


class NormeServiceViewSet(viewsets.ModelViewSet):
    queryset = NormeService.objects.all()
    serializer_class = NormeServiceSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        type_structure = self.request.query_params.get("type_structure")
        # ✅ FIX: Only filter if type_structure is not null/None
        if type_structure and type_structure != "null":
            try:
                qs = qs.filter(type_structure_id=int(type_structure))
            except (ValueError, TypeError):
                pass
        return qs.order_by("nom_service")


class NormeMaterielViewSet(viewsets.ModelViewSet):
    queryset = NormeMateriel.objects.all()
    serializer_class = NormeMaterielSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        type_structure = self.request.query_params.get("type_structure")
        # ✅ FIX: Only filter if type_structure is not null/None
        if type_structure and type_structure != "null":
            try:
                qs = qs.filter(type_structure_id=int(type_structure))
            except (ValueError, TypeError):
                pass
        return qs.order_by("nom_materiel")
    
from django.db.models import Q
from rest_framework import viewsets


# class StructureSanteViewSet(viewsets.ModelViewSet):
#     queryset = StructureSante.objects.select_related(
#         "type_structure", "wilaya_fk", "moughataa_fk", "commune_fk"
#     ).all()
#     serializer_class = StructureSanteSerializer
#     parser_classes = [MultiPartParser]
#     permission_classes = [permissions.IsAuthenticated ,CustomModelPermissions, FOSARolePermission]

#     def get_queryset(self):
#         qs = super().get_queryset()

#         # accept BOTH names (so React can send wilaya or wilaya_fk)
#         wilaya = self.request.query_params.get("wilaya_fk") or self.request.query_params.get("wilaya")
#         moughataa = self.request.query_params.get("moughataa_fk") or self.request.query_params.get("moughataa")
#         type_structure = self.request.query_params.get("type_structure")
#         q = self.request.query_params.get("q")

        
#         if wilaya:
#             qs = qs.filter(
#                 wilaya_fk_id=wilaya if str(wilaya).isdigit() else None
#             ) if str(wilaya).isdigit() else qs.filter(wilaya_fk__nom__iexact=wilaya)

#         if moughataa:
#             qs = qs.filter(
#                 moughataa_fk_id=moughataa if str(moughataa).isdigit() else None
#             ) if str(moughataa).isdigit() else qs.filter(moughataa_fk__nom__iexact=moughataa)

       
#         if type_structure:
#             qs = qs.filter(type_structure_id=type_structure)

#         if q:
#             qs = qs.filter(
#                 Q(code__icontains=q) |
#                 Q(structure__icontains=q) |
#                 Q(wilaya__icontains=q) |
#                 Q(moughataa__icontains=q) |
#                 Q(commune__icontains=q) |
#                 Q(responsable__icontains=q)
#             )

#         return qs.order_by("wilaya_fk__nom", "moughataa_fk__nom", "structure")
#     # -----------------------------
#     # Tes actions existantes
#     # -----------------------------
#     @action(detail=True, methods=["get"])
#     def personnels(self, request, pk=None):
#         structure = self.get_object()
#         qs = structure.personnels.all().order_by("intitule_poste")
#         return Response(PersonnelStructureSerializer(qs, many=True).data)

#     @action(detail=True, methods=["post"], url_path="personnels/upsert")
#     def personnels_upsert(self, request, pk=None):
#         structure = self.get_object()
#         items = request.data if isinstance(request.data, list) else [request.data]

#         saved = []
#         for it in items:
#             intitule = it.get("intitule_poste")
#             nombre = it.get("nombre_reel", 0)
#             if not intitule:
#                 return Response({"detail": "intitule_poste manquant"}, status=400)

#             obj, _ = PersonnelStructure.objects.update_or_create(
#                 structure=structure,
#                 intitule_poste=intitule,
#                 defaults={"nombre_reel": nombre},
#             )
#             saved.append(PersonnelStructureSerializer(obj).data)

#         return Response(saved, status=status.HTTP_200_OK)

#     @action(detail=True, methods=["get"])
#     def services(self, request, pk=None):
#         structure = self.get_object()
#         qs = structure.services.all().order_by("nom_service")
#         return Response(ServiceStructureSerializer(qs, many=True).data)

#     @action(detail=True, methods=["post"], url_path="services/upsert")
#     def services_upsert(self, request, pk=None):
#         structure = self.get_object()
#         items = request.data if isinstance(request.data, list) else [request.data]

#         saved = []
#         for it in items:
#             nom = it.get("nom_service")
#             dispo = bool(it.get("disponible", False))
#             if not nom:
#                 return Response({"detail": "nom_service manquant"}, status=400)

#             obj, _ = ServiceStructure.objects.update_or_create(
#                 structure=structure,
#                 nom_service=nom,
#                 defaults={"disponible": dispo},
#             )
#             saved.append(ServiceStructureSerializer(obj).data)

#         return Response(saved, status=status.HTTP_200_OK)

#     @action(detail=True, methods=["get"])
#     def materiels(self, request, pk=None):
#         structure = self.get_object()
#         qs = structure.materiels.all().order_by("nom_materiel")
#         return Response(MaterielStructureSerializer(qs, many=True).data)

#     @action(detail=True, methods=["post"], url_path="materiels/upsert")
#     def materiels_upsert(self, request, pk=None):
#         structure = self.get_object()
#         items = request.data if isinstance(request.data, list) else [request.data]

#         saved = []
#         for it in items:
#             nom = it.get("nom_materiel")
#             qte = it.get("quantite_reelle", 0)
#             if not nom:
#                 return Response({"detail": "nom_materiel manquant"}, status=400)

#             obj, _ = MaterielStructure.objects.update_or_create(
#                 structure=structure,
#                 nom_materiel=nom,
#                 defaults={"quantite_reelle": qte},
#             )
#             saved.append(MaterielStructureSerializer(obj).data)

#         return Response(saved, status=status.HTTP_200_OK)